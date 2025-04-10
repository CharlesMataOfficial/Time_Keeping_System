import json
import os
from datetime import date, datetime, time, timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import login, logout
from django.urls import reverse
from .models import (
    CustomUser,
    TimeEntry,
    Announcement,
    AdminLog,
    Leave,
)
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook

from .utils import (
    COMPANY_CHOICES,
    DEPARTMENT_CHOICES,
    get_company_logo,
    log_admin_action,
)

@never_cache
def login_view(request):
    """
    Handles user login.

    Authenticates users based on employee ID and PIN. Redirects staff/superusers
    to the admin page and guards to the user page.

    Args:
        request: The HTTP request object.

    Returns:
        HttpResponse: Renders the login page or redirects to the appropriate page
                      after successful login.
    """
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        pin = request.POST.get("pin")

        try:
            # First check if user exists and is active
            user = CustomUser.objects.get(employee_id=employee_id)

            if not user.is_active:
                return render(
                    request, "login_page.html", {"error": "This account is inactive"}
                )

            # Try to authenticate
            auth_result = CustomUser.authenticate_by_pin(employee_id, pin)

            if auth_result:  # Successful login
                user = (
                    auth_result
                    if isinstance(auth_result, CustomUser)
                    else auth_result["user"]
                )
                login(request, user)
                log_admin_action(request, "login", f"Successfully logged in")

                if user.is_guard:
                    return redirect("user_page")
                elif user.is_staff or user.is_superuser:
                    return redirect("custom_admin_page")
                else:
                    return render(
                        request,
                        "login_page.html",
                        {"error": "You do not have permission to log in"},
                    )
            else:
                return render(request, "login_page.html", {"error": "Incorrect PIN"})

        except CustomUser.DoesNotExist:
            return render(
                request, "login_page.html", {"error": "Employee ID not found"}
            )

    return render(request, "login_page.html")


@login_required
def user_page(request):
    """
    Renders the user page for guards.

    Retrieves today's time entries and renders the user page.

    Args:
        request: The HTTP request object.

    Returns:
        HttpResponse: Renders the user page with time entry data.
    """
    # Force a refresh from the DB so that the latest value of is_guard is loaded
    request.user.refresh_from_db()

    if not request.user.is_guard:
        messages.error(
            request, "Access denied. You do not have permission to access this page."
        )
        return redirect("custom_admin_page")

    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    todays_entries = TimeEntry.objects.filter(
        time_in__gte=today_start, time_in__lt=today_end
    ).order_by("-last_modified")

    return render(
        request,
        "user_page.html",
        {
            "all_entries": todays_entries,
            "partner_logo": "default_logo.png",
            "user_company": "",
        },
    )


def logout_view(request):
    """
    Logs out the user.

    Logs the logout action and redirects to the login page.

    Args:
        request: The HTTP request object.

    Returns:
        HttpResponseRedirect: Redirects to the login page.
    """
    if request.user.is_authenticated:
        log_admin_action(request, 'logout', f'User logged out')
    logout(request)
    return redirect("login")


@require_POST
def clock_in_view(request):
    """
    Handles clock-in requests.

    Authenticates the user, captures an image (if available), and creates a new
    time entry.

    Args:
        request: The HTTP request object containing employee ID, PIN, and
                 optionally, a new PIN for first-time login.

    Returns:
        JsonResponse: Indicates success or failure, along with user and time
                      entry data.
    """
    data = json.loads(request.body)
    employee_id = data.get("employee_id")
    pin = data.get("pin")
    new_pin = data.get("new_pin")
    image_path = data.get("image_path")
    first_login_check = data.get("first_login_check", False)

    auth_result = CustomUser.authenticate_by_pin(employee_id, pin)

    # Handle first login cases
    if isinstance(auth_result, dict) and auth_result.get("status") == "first_login":
        if new_pin:
            user = auth_result["user"]
            user.pin = new_pin
            user.if_first_login = False
            user.save()
            return JsonResponse({"success": True, "message": "PIN updated successfully"})
        else:
            return JsonResponse({
                "success": False,
                "error": "first_login",
                "message": "Please set your new PIN"
            })

    # For first login check only
    if first_login_check:
        if isinstance(auth_result, dict) and auth_result.get("status") == "first_login":
            return JsonResponse({
                "success": False,
                "error": "first_login",
                "message": "Please set your new PIN"
            })
        return JsonResponse({"success": True})

    # If authentication failed
    if not auth_result:
        try:
            CustomUser.objects.get(employee_id=employee_id)
            error_message = "Incorrect PIN"
        except CustomUser.DoesNotExist:
            error_message = "Employee ID not found"
        return JsonResponse({"success": False, "error": error_message})

    # Authentication passed: get the user
    user = auth_result

    # --- Prevent Multiple Clock-Ins in the Same Day ---
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    # Check if there is already a clock in entry for today for this user
    if TimeEntry.objects.filter(user=user, time_in__gte=today_start).exists():
        return JsonResponse({"success": False, "error": "You have already clocked in today"})

    # --- Check for Previous Day Entry Without Clock-Out ---
    yesterday_start = today_start - timedelta(days=1)
    yesterday_end = today_start
    yesterday_entry = TimeEntry.objects.filter(
        user=user,
        time_in__gte=yesterday_start,
        time_in__lt=yesterday_end,
        time_out__isnull=True
    ).first()
    warning_message = None
    if yesterday_entry:
        warning_message = "Clock in successful! However, you forgot to clock out yesterday."

    # --- Proceed with Creating a New Clock-In Entry ---
    entry = TimeEntry.clock_in(user)
    if image_path:
        entry.image_path = image_path
        entry.save()

    # Get the company logo using the utility function
    user_company = user.company.name if user.company else ""
    company_logo = get_company_logo(user_company)

    # Fetch updated attendance list for today
    today_end = today_start + timedelta(days=1)
    todays_entries = TimeEntry.objects.filter(
        time_in__gte=today_start, time_in__lt=today_end
    ).order_by("-last_modified")

    attendance_list = [
        {
            "employee_id": entry.user.employee_id,
            "first_name": entry.user.first_name,
            "surname": entry.user.surname,
            "company": entry.user.company.name if entry.user.company else "",
            "time_in": entry.time_in.strftime("%I:%M %p"),
            "time_out": (entry.time_out.strftime("%I:%M %p") if entry.time_out else None),
            "image_path": entry.image_path,
        }
        for entry in todays_entries
    ]

    response = {
        "success": True,
        "employee_id": user.employee_id,
        "first_name": user.first_name,
        "surname": user.surname,
        "company": user.company.name if user.company else "",
        "time_in": entry.time_in.strftime("%I:%M %p"),
        "time_out": None,
        "image_path": entry.image_path,
        "new_logo": company_logo,
        "attendance_list": attendance_list,
    }
    if warning_message:
        response["warning"] = warning_message

    return JsonResponse(response)


@require_POST
def clock_out_view(request):
    """
    Handles clock-out requests.

    Authenticates the user and updates the latest time entry with a clock-out time.

    Args:
        request: The HTTP request object containing employee ID and PIN.

    Returns:
        JsonResponse: Indicates success or failure, along with user and time
                      entry data.
    """
    data = json.loads(request.body)
    employee_id = data.get("employee_id")
    pin = data.get("pin")

    user = CustomUser.authenticate_by_pin(employee_id, pin)

    if user:
        try:
            now = timezone.now()
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            today_end = today_start + timedelta(days=1)

            open_entry = TimeEntry.objects.filter(
                user=user,
                time_out__isnull=True,
                time_in__gte=today_start,
                time_in__lt=today_end,
            ).latest("time_in")

            open_entry.clock_out()

            time_in_formatted = open_entry.time_in.strftime("%I:%M %p")
            time_out_formatted = open_entry.time_out.strftime("%I:%M %p")

            # Handle None company value
            user_company = user.company.name if user.company else ""
            company_logo = get_company_logo(user_company)

            # Fetch updated attendance list
            todays_entries = TimeEntry.objects.filter(
                time_in__gte=today_start, time_in__lt=today_end
            ).order_by("-last_modified")

            attendance_list = [
                {
                    "employee_id": entry.user.employee_id,
                    "first_name": entry.user.first_name,
                    "surname": entry.user.surname,
                    "company": entry.user.company.name if entry.user.company else "",
                    "time_in": entry.time_in.strftime("%I:%M %p"),
                    "time_out": (
                        entry.time_out.strftime("%I:%M %p") if entry.time_out else None
                    ),
                    "image_path": entry.image_path,
                }
                for entry in todays_entries
            ]

            return JsonResponse(
                {
                    "success": True,
                    "employee_id": user.employee_id,
                    "first_name": user.first_name or "",
                    "surname": user.surname or "",
                    "company": user.company.name if user.company else "",
                    "time_in": time_in_formatted,
                    "time_out": time_out_formatted,
                    "new_logo": company_logo,
                    "attendance_list": attendance_list,
                }
            )
        except TimeEntry.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "No active clock in found."}
            )
        except Exception as e:
            return JsonResponse(
                {"success": False, "error": f"An error occurred: {str(e)}"}
            )
    else:
        try:
            CustomUser.objects.get(employee_id=employee_id)
            error_message = "Incorrect PIN"
        except CustomUser.DoesNotExist:
            error_message = "Employee ID not found"
        return JsonResponse({"success": False, "error": error_message})


@require_GET
@login_required
def get_todays_entries(request):
    """
    Retrieves today's time entries.

    Retrieves all time entries for the current day and returns them as a JSON
    response.

    Args:
        request: The HTTP request object.

    Returns:
        JsonResponse: A JSON response containing a list of time entries.
    """
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)

    # Make sure entries are ordered by last_modified in descending order
    entries = TimeEntry.objects.filter(
        time_in__gte=today_start, time_in__lt=today_end
    ).order_by(
        "-last_modified"
    )  # This is correct

    entries_data = []
    for entry in entries:
        entries_data.append(
            {
                "employee_id": entry.user.employee_id,
                "first_name": entry.user.first_name,
                "surname": entry.user.surname,
                "company": entry.user.company.name if entry.user.company else "",
                "time_in": entry.time_in.strftime("%I:%M %p"),
                "time_out": (
                    entry.time_out.strftime("%I:%M %p") if entry.time_out else None
                ),
            }
        )

    return JsonResponse({"entries": entries_data})


@require_POST
def upload_image(request):
    """
    Uploads an image associated with a time entry.

    Saves the uploaded image to the server and returns the file path.

    Args:
        request: The HTTP request object containing the image and employee ID.

    Returns:
        JsonResponse: Indicates success or failure, along with the file path.
    """
    image_data = request.FILES.get("image")
    employee_id = request.POST.get("employee_id")

    if image_data:
        try:
            user = CustomUser.objects.get(employee_id=employee_id)
        except CustomUser.DoesNotExist:
            return JsonResponse({"success": False, "error": "Employee ID not found"})

        # Get the current date
        now = datetime.now()
        year = now.strftime("%Y")
        month = now.strftime("%m")
        day = now.strftime("%d")
        timestamp = now.strftime("%H%M%S")

        # Create directories based on the current date
        directory = os.path.join("attendance_images", year, month, day)
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Create a unique file name
        file_name = (
            f"{timestamp}_{user.employee_id}_{user.surname}{user.first_name}.jpg"
        )
        file_path = os.path.join(directory, file_name)

        # Save the file
        file_path = default_storage.save(file_path, ContentFile(image_data.read()))
        return JsonResponse({"success": True, "file_path": file_path})
    return JsonResponse({"success": False, "error": "No image uploaded"})


@csrf_exempt
def announcements_list_create(request):
    """
    API endpoint for listing and creating announcements.

    GET: Returns a list of all announcements (JSON).
    POST: Creates a new announcement (expects JSON body { content: "..."} ).

    Args:
        request: The HTTP request object.

    Returns:
        JsonResponse: A JSON response containing announcement data or a
                      success/error message.
    """
    if request.method == "GET":
        announcements = Announcement.objects.all().order_by("-created_at")
        data = [
            {
                "id": ann.id,
                "content": ann.content,
                "created_at": ann.created_at.isoformat(),
                "is_posted": ann.is_posted,
            }
            for ann in announcements
        ]
        return JsonResponse(data, safe=False)

    elif request.method == "POST":
        try:
            body = json.loads(request.body)
            content = body.get("content", "")
            announcement = Announcement.objects.create(content=content)
            return JsonResponse(
                {"message": "Announcement created", "id": announcement.id}
            )
        except:
            return HttpResponseBadRequest("Invalid data")

    return HttpResponseBadRequest("Unsupported method")


@csrf_exempt
def announcement_detail(request, pk):
    """
    API endpoint for retrieving a single announcement.

    GET: Returns details of a single announcement by ID.

    Args:
        request: The HTTP request object.
        pk: The primary key of the announcement.

    Returns:
        JsonResponse: A JSON response containing the announcement data.
    """
    announcement = get_object_or_404(Announcement, pk=pk)

    if request.method == "GET":
        data = {
            "id": announcement.id,
            "content": announcement.content,
            "created_at": announcement.created_at.isoformat(),
            "is_posted": announcement.is_posted,
        }
        return JsonResponse(data)

    return HttpResponseBadRequest("Unsupported method")


@csrf_exempt
def announcement_delete(request, pk):
    """
    API endpoint for deleting an announcement.

    DELETE: Deletes an announcement by ID.

    Args:
        request: The HTTP request object.
        pk: The primary key of the announcement.

    Returns:
        JsonResponse: A JSON response indicating success.
    """
    if request.method == "DELETE":
        announcement = get_object_or_404(Announcement, pk=pk)
        announcement.delete()
        return JsonResponse({"message": "Announcement deleted"})
    return HttpResponseBadRequest("Unsupported method")


@csrf_exempt
def announcement_post(request, pk):
    """
    API endpoint for posting an announcement.

    POST: Marks an announcement as posted (is_posted = True).

    Args:
        request: The HTTP request object.
        pk: The primary key of the announcement.

    Returns:
        JsonResponse: A JSON response indicating success.
    """
    if request.method == "POST":
        announcement = get_object_or_404(Announcement, pk=pk)
        announcement.is_posted = True
        announcement.save()

        log_admin_action(request, 'announcement_post', f'Posted announcement: "{announcement.content[:30]}..."')
        return JsonResponse({"message": "Announcement posted"})
    return HttpResponseBadRequest("Unsupported method")


@csrf_exempt
def posted_announcements_list(request):
    """
    API endpoint for listing posted announcements.

    GET: Returns a list of posted announcements (is_posted=True).

    Args:
        request: The HTTP request object.

    Returns:
        JsonResponse: A JSON response containing a list of posted
                      announcements.
    """
    if request.method == "GET":
        # Filter to only posted announcements
        announcements = Announcement.objects.filter(is_posted=True).order_by(
            "-created_at"
        )
        data = [
            {
                "id": ann.id,
                "content": ann.content,
                "created_at": ann.created_at.isoformat(),
                "is_posted": ann.is_posted,
            }
            for ann in announcements
        ]
        return JsonResponse(data, safe=False)

    return HttpResponseBadRequest("Unsupported method")


@login_required
def custom_admin_page(request):
    """
    Renders the custom admin page.

    Redirects non-staff/non-superusers to the user page.

    Args:
        request: The HTTP request object.

    Returns:
        HttpResponse: Renders the custom admin page.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect("user_page")

    log_admin_action(request, "navigation", "Accessed the admin dashboard")
    return render(request, "custom_admin_page.html")


@login_required
def superadmin_redirect(request):
    """
    Redirects superusers to the Django admin page.

    Args:
        request: The HTTP request object.

    Returns:
        HttpResponseRedirect: Redirects to the Django admin page.
    """
    if request.user.is_superuser:
        log_admin_action(request, "navigation", "Accessed the superadmin page")
        return redirect(reverse("admin:index"))
    else:
        messages.error(
            request, "You do not have permission to access the super admin page."
        )
        return redirect("custom_admin_page")


def get_special_dates(request):
    """
    Retrieves special dates (birthdays and hiring anniversaries).

    Retrieves users with birthdays and hiring anniversaries for the current day.

    Args:
        request: The HTTP request object.

    Returns:
        JsonResponse: A JSON response containing lists of birthdays and
                      milestones.
    """
    today = timezone.now().date()

    # Get users with birthdays today based on 'birth_date'
    birthday_users = list(
        CustomUser.objects.filter(
            birth_date__month=today.month, birth_date__day=today.day
        ).values("employee_id", "first_name", "surname")
    )

    # Get users with hiring anniversaries today based on 'date_hired'
    milestone_users = []
    for user in CustomUser.objects.filter(
        date_hired__month=today.month, date_hired__day=today.day
    ):
        years = today.year - user.date_hired.year
        if years >= 1:
            milestone_users.append(
                {
                    "employee_id": user.employee_id,
                    "first_name": user.first_name,
                    "surname": user.surname,
                    "years": years,
                }
            )

    return JsonResponse({"birthdays": birthday_users, "milestones": milestone_users})


@login_required
def attendance_list_json(request):
    """
    API endpoint for retrieving attendance data.

    Retrieves attendance data based on various filter parameters and returns it
    as a JSON response.

    Args:
        request: The HTTP request object containing filter parameters.

    Returns:
        JsonResponse: A JSON response containing a list of attendance records.
    """
    # Get filter parameters
    attendance_type = request.GET.get("attendance_type", "time-log")
    company_code = request.GET.get("attendance_company", "all")
    department_code = request.GET.get("attendance_department", "all")
    search_query = request.GET.get("search", "").strip()

    # Pagination parameters
    try:
        page = int(request.GET.get("page", 1))
        limit = int(request.GET.get("limit", 50))
    except ValueError:
        page = 1
        limit = 50

    # Base queries with appropriate filtering
    if attendance_type == "time-log":
        # Only include time entries for today
        today = date.today()
        qs = (
            TimeEntry.objects.select_related("user", "user__company", "user__position")
            .filter(time_in__date=today, user__is_active=True)
            .order_by("-last_modified")
        )

        # Apply company filter
        if company_code != "all":
            companies_to_filter = []
            # Check if company_code is directly in COMPANY_CHOICES
            if company_code in COMPANY_CHOICES:
                companies_to_filter = COMPANY_CHOICES[company_code]
            else:
                # Check if company_code is an alias in any tuple
                for key, names_tuple in COMPANY_CHOICES.items():
                    if company_code in names_tuple:
                        companies_to_filter = names_tuple
                        break

            if companies_to_filter:
                query = Q()
                for company_name in companies_to_filter:
                    query |= Q(user__company__name__iexact=company_name)
                qs = qs.filter(query)
            else:
                qs = qs.filter(user__company__name__iexact=company_code)

        # Apply department filter
        if department_code != "all":
            if department_code in DEPARTMENT_CHOICES:
                dept_name = DEPARTMENT_CHOICES[department_code]
                qs = qs.filter(user__position__name=dept_name)
            else:
                qs = qs.filter(user__position__name=department_code)

        # Apply search filter
        if search_query:
            # For more exact matching, try to match the complete name first
            complete_name_query = (
                Q(user__first_name__icontains=search_query)
                | Q(user__surname__icontains=search_query)
                | Q(user__first_name__iexact=search_query)
                | Q(user__surname__iexact=search_query)
            )

            # Then try concatenated name (first_name + surname)
            name_parts = search_query.split()
            if len(name_parts) > 1:
                # If we have multiple terms, try to match them exactly in order
                first_name_candidates = [
                    " ".join(name_parts[:i]) for i in range(1, len(name_parts))
                ]
                surname_candidates = [
                    " ".join(name_parts[i:]) for i in range(1, len(name_parts))
                ]

                for first in first_name_candidates:
                    for last in surname_candidates:
                        if first and last:  # Ensure we don't have empty strings
                            complete_name_query |= Q(
                                user__first_name__iexact=first
                            ) & Q(user__surname__iexact=last)

            qs = qs.filter(complete_name_query)

        # Get total count before pagination
        total_count = qs.count()

        # Apply pagination
        start = (page - 1) * limit
        end = page * limit
        paginated_qs = qs[start:end]

        # Format time-log data
        attendance_list = [
            {
                "employee_id": entry.user.employee_id,
                "name": f"{entry.user.first_name} {entry.user.surname}",
                "time_in": entry.time_in.strftime("%Y-%m-%d %H:%M:%S"),
                "time_out": (
                    entry.time_out.strftime("%Y-%m-%d %H:%M:%S")
                    if entry.time_out
                    else ""
                ),
                "hours_worked": entry.hours_worked,
            }
            for entry in paginated_qs
        ]

    elif attendance_type in ["users-active", "users-inactive"]:
        if attendance_type == "users-active":
            qs = CustomUser.objects.filter(is_active=True).distinct()
        else:  # users-inactive
            qs = CustomUser.objects.filter(is_active=False).distinct()

        # Apply company filter
        if company_code != "all":
            companies_to_filter = []
            if company_code in COMPANY_CHOICES:
                companies_to_filter = COMPANY_CHOICES[company_code]
            else:
                for key, names_tuple in COMPANY_CHOICES.items():
                    if company_code in names_tuple:
                        companies_to_filter = names_tuple
                        break

            if companies_to_filter:
                query = Q()
                for company_name in companies_to_filter:
                    query |= Q(company__name__iexact=company_name)
                qs = qs.filter(query)
            else:
                qs = qs.filter(company__name__iexact=company_code)

        # Apply department filter
        if department_code != "all":
            qs = qs.filter(position__name=department_code)

        # Apply search filter
        if search_query:
            # For more exact matching, try to match the complete name first
            complete_name_query = (
                Q(first_name__icontains=search_query)
                | Q(surname__icontains=search_query)
                | Q(first_name__iexact=search_query)
                | Q(surname__iexact=search_query)
            )

            # Then try concatenated name (first_name + surname)
            name_parts = search_query.split()
            if len(name_parts) > 1:
                # If we have multiple terms, try to match them exactly in order
                first_name_candidates = [
                    " ".join(name_parts[:i]) for i in range(1, len(name_parts))
                ]
                surname_candidates = [
                    " ".join(name_parts[i:]) for i in range(1, len(name_parts))
                ]

                for first in first_name_candidates:
                    for last in surname_candidates:
                        if first and last:  # Ensure we don't have empty strings
                            complete_name_query |= Q(first_name__iexact=first) & Q(
                                surname__iexact=last
                            )

            qs = qs.filter(complete_name_query)

        # Get total count before pagination
        total_count = qs.count()

        # Apply pagination
        start = (page - 1) * limit
        end = page * limit
        paginated_qs = qs[start:end]

        # Format user data
        attendance_list = [
            {
                "employee_id": user.employee_id,
                "name": f"{user.first_name} {user.surname}",
            }
            for user in paginated_qs
        ]
    else:
        attendance_list = []
        total_count = 0

    return JsonResponse({
        "attendance_list": attendance_list,
        "attendance_type": attendance_type,
        "total": total_count
    })


@login_required
@require_GET
def dashboard_data(request):
    """
    API endpoint for retrieving dashboard data.

    Retrieves data for the admin dashboard, including today's time entries,
    top late employees, and top early birds.

    Args:
        request: The HTTP request object.

    Returns:
        JsonResponse: A JSON response containing dashboard data.
    """
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)

    # Get all entries for today
    todays_entries = TimeEntry.objects.filter(
        time_in__gte=today_start, time_in__lt=today_end
    ).select_related("user", "user__company", "user__schedule_group")

    processed_entries = []
    late_count = 0

    for entry in todays_entries:
        user = entry.user
        time_in_local = entry.time_in

        # Get user name or default to ID
        first_name = user.first_name or ""
        surname = user.surname or ""
        full_name = f"{first_name} {surname}".strip()
        if not full_name:
            full_name = f"User {user.employee_id}"

        # Use the stored minutes_late value
        if entry.is_late:
            late_count += 1

        processed_entries.append(
            {
                "employee_id": user.employee_id,
                "name": full_name,
                "company": user.company.name if user.company else "",
                "time_in": time_in_local.strftime("%I:%M %p"),
                "time_out": (
                    entry.time_out.strftime("%I:%M %p") if entry.time_out else None
                ),
                "minutes_diff": entry.minutes_late,  # Use the stored value
                "is_late": entry.is_late,
            }
        )

    # Sort entries - late ones by how late they are (descending)
    late_entries = sorted(
        [e for e in processed_entries if e["is_late"]],
        key=lambda x: x["minutes_diff"],
        reverse=True,
    )[:5]

    # Sort entries - early ones by how early they are (most early first)
    early_entries = sorted(
        [e for e in processed_entries if e["minutes_diff"] < 0],  # Only include truly early entries
        key=lambda x: x["minutes_diff"],  # Ascending order for negative numbers puts most negative (most early) first
    )[:5]

    return JsonResponse(
        {
            "today_entries": processed_entries,
            "top_late": late_entries,
            "top_early": early_entries,
            "late_count": late_count,
        }
    )


@login_required
@require_GET
def get_logs(request):
    """
    API endpoint for retrieving admin logs.

    Retrieves admin logs based on various filter parameters and returns them as a
    JSON response.

    Args:
        request: The HTTP request object containing filter parameters.

    Returns:
        JsonResponse: A JSON response containing a list of admin logs.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"error": "Permission denied"}, status=403)

    # Get filter parameters
    search_query = request.GET.get("search", "").strip()
    action_filter = request.GET.get("action", "")
    date_range = request.GET.get("date_range", "")

    # Pagination parameters
    try:
        page = int(request.GET.get("page", 1))
        limit = int(request.GET.get("limit", 50))
    except ValueError:
        page = 1
        limit = 50

    # Base query
    logs_query = AdminLog.objects.all()

    # Apply filters
    if search_query:
        logs_query = logs_query.filter(
            Q(description__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__surname__icontains=search_query) |
            Q(user__employee_id__icontains=search_query)
        )

    if action_filter:
        logs_query = logs_query.filter(action=action_filter)

    # Date range filtering
    now = timezone.now()
    if date_range == "today":
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        logs_query = logs_query.filter(timestamp__gte=today_start)
    elif date_range == "week":
        week_start = now - timedelta(days=now.weekday())
        week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        logs_query = logs_query.filter(timestamp__gte=week_start)
    elif date_range == "month":
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        logs_query = logs_query.filter(timestamp__gte=month_start)

    # Calculate pagination
    start = (page - 1) * limit
    end = page * limit

    # Get paginated logs
    logs = logs_query[start:end]

    log_data = [
        {
            "user": f"{log.user.first_name} {log.user.surname}",
            "employee_id": log.user.employee_id,
            "action": log.get_action_display(),
            "description": log.description,
            "timestamp": log.timestamp.strftime("%Y-%m-%d %I:%M %p"),
            "ip_address": log.ip_address or "Unknown",
        }
        for log in logs
    ]

    return JsonResponse({"logs": log_data, "total": logs_query.count()})


@require_GET
def export_time_entries_by_date(request):
    """
    API endpoint for exporting time entries to Excel by date.

    Exports time entries for a specific date to an Excel file.

    Args:
        request: The HTTP request object containing the date parameter.

    Returns:
        HttpResponse: An Excel file containing the time entries.
    """
    date_str = request.GET.get("date")
    if not date_str:
        return HttpResponse("Date parameter is required.", status=400)

    selected_date = parse_date(date_str)
    if not selected_date:
        return HttpResponse("Invalid date format.", status=400)

    start_datetime = datetime.combine(selected_date, time.min)
    end_datetime = datetime.combine(selected_date, time.max)

    qs = TimeEntry.objects.filter(
        time_in__gte=start_datetime,
        time_in__lte=end_datetime
    ).order_by("time_in")

    employee_id = request.GET.get("employee_id")
    if employee_id:
        # Remove the int conversion to match the behavior of your date range export
        qs = qs.filter(user__employee_id=employee_id)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Entries"
    headers = [
        "ID",
        "Employee ID",
        "First Name",
        "Surname",
        "Company",
        "Date",
        "Time In",
        "Time Out",
        "Hours Worked",
        "Is Late"
    ]
    ws.append(headers)
    for entry in qs:
        row = [
            entry.id,
            entry.user.employee_id,
            entry.user.first_name,
            entry.user.surname,
            entry.user.company.name if entry.user.company else "",
            entry.time_in.strftime("%Y-%m-%d"),
            entry.time_in.strftime("%H:%M:%S"),
            entry.time_out.strftime("%H:%M:%S") if entry.time_out else "",
            entry.hours_worked,
            "Yes" if entry.is_late else "No"
        ]
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"time_entries_{date_str}.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"

    log_admin_action(request, "Excel Export", f"Exported {filename}")
    log_admin_action(request, "excel_export", f"Exported {filename}")
    return response

@require_GET
def export_time_entries_range(request):
    """
    API endpoint for exporting time entries to Excel by date range.

    Exports time entries for a specific date range to an Excel file.

    Args:
        request: The HTTP request object containing the start and end date parameters.

    Returns:
        HttpResponse: An Excel file containing the time entries.
    """
    date_start_str = request.GET.get("date_start")
    date_end_str = request.GET.get("date_end")
    if not date_start_str or not date_end_str:
        return HttpResponse("Start date and End date parameters are required.", status=400)

    date_start = parse_date(date_start_str)
    date_end = parse_date(date_end_str)
    if not date_start or not date_end:
        return HttpResponse("Invalid date format.", status=400)

    start_datetime = datetime.combine(date_start, time.min)
    end_datetime = datetime.combine(date_end, time.max)

    qs = TimeEntry.objects.filter(
        time_in__gte=start_datetime,
        time_in__lte=end_datetime
    )

    # Exclude dates if provided
    excluded_dates = request.GET.getlist("exclude_date")
    if excluded_dates:
        qs = qs.exclude(time_in__date__in=excluded_dates)

    # Optional employee id filter
    employee_id = request.GET.get("employee_id")
    if employee_id:
        qs = qs.filter(user__employee_id=employee_id)

    qs = qs.order_by("time_in")

    # Create Excel workbook and sheet (code unchanged)
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Entries"
    headers = [
        "ID",
        "Employee ID",
        "First Name",
        "Surname",
        "Company",
        "Date",
        "Time In",
        "Time Out",
        "Hours Worked",
        "Is Late"
    ]
    ws.append(headers)
    for entry in qs:
        row = [
            entry.id,
            entry.user.employee_id,
            entry.user.first_name,
            entry.user.surname,
            entry.user.company.name if entry.user.company else "",
            entry.time_in.strftime("%Y-%m-%d"),
            entry.time_in.strftime("%H:%M:%S"),
            entry.time_out.strftime("%H:%M:%S") if entry.time_out else "",
            entry.hours_worked,
            "Yes" if entry.is_late else "No"
        ]
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"time_entries_range_{date_start_str}_to_{date_end_str}.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=time_entries_{employee_id}.xlsx"
    log_admin_action(request, "Excel Export", f"Exported time_entries_{employee_id}.xlsx")
    return response

@login_required
def get_pending_leaves(request):
    """
    API endpoint for retrieving pending leave requests.

    Retrieves pending leave requests based on user role (HR or manager).

    Args:
        request: The HTTP request object.

    Returns:
        JsonResponse: A JSON response containing a list of pending leave requests.
    """
    user = request.user

    try:
        # Determine which leaves to show based on user role
        if user.is_hr:
            # HR sees leaves that managers have approved
            leaves = Leave.objects.filter(status='APPROVED_BY_MANAGER')
        else:
            # Regular users (managers) see pending leaves from their team
            leaves = Leave.objects.filter(employee__manager=user, status='PENDING')

        # Format the leave data
        leave_data = []
        for leave in leaves:
            leave_data.append({
                'id': leave.id,
                'employee_name': f"{leave.employee.first_name} {leave.employee.surname}",
                'start_date': leave.start_date.strftime('%Y-%m-%d'),
                'end_date': leave.end_date.strftime('%Y-%m-%d'),
                'duration': leave.get_duration(),
                'leave_type': leave.leave_type.name if leave.leave_type else 'N/A',
                'reason': leave.reason
            })

        return JsonResponse({'leaves': leave_data})
    except Exception as e:
        # Log the error
        import traceback
        print(f"Error in get_pending_leaves: {e}")
        print(traceback.format_exc())
        # Return empty leaves array with error status
        return JsonResponse({'leaves': [], 'error': str(e)})

@login_required
@require_POST
def process_leave(request):
    """
    API endpoint for processing leave requests.

    Processes leave requests (approve or reject) based on user role (HR or manager).

    Args:
        request: The HTTP request object containing leave ID and action.

    Returns:
        JsonResponse: A JSON response indicating success or failure.
    """
    leave_id = request.POST.get('leave_id')
    action = request.POST.get('action')  # 'approve' or 'reject'
    rejection_reason = request.POST.get('rejection_reason', '')

    leave = get_object_or_404(Leave, id=leave_id)
    user = request.user

    # Verify permission to process this leave
    if user.is_hr and leave.status == 'APPROVED_BY_MANAGER':
        if action == 'approve':
            leave.status = 'APPROVED_BY_HR'
            # We'll leave credits tracking for later
        else:
            leave.status = 'REJECTED_BY_HR'
            leave.rejection_reason = rejection_reason

    elif leave.employee.manager == user and leave.status == 'PENDING':
        if action == 'approve':
            leave.status = 'APPROVED_BY_MANAGER'
        else:
            leave.status = 'REJECTED_BY_MANAGER'
            leave.rejection_reason = rejection_reason
    else:
        return JsonResponse({'success': False, 'message': 'Not authorized to process this leave'})

    leave.save()

    # Log the action
    log_admin_action(request, 'leave_approval', f"Leave request {action}d for {leave.employee.employee_id}")

    return JsonResponse({'success': True})

@require_GET
def export_time_entries_by_date(request):
    """
    API endpoint for exporting time entries to Excel by date.

    Exports time entries for a specific date to an Excel file.

    Args:
        request: The HTTP request object containing the date parameter.

    Returns:
        HttpResponse: An Excel file containing the time entries.
    """
    # Get the date parameter from the request (e.g. from your date picker)
    date_str = request.GET.get("date")
    if not date_str:
        return HttpResponse("Date parameter is required.", status=400)

    selected_date = parse_date(date_str)
    if not selected_date:
        return HttpResponse("Invalid date format.", status=400)

    # Create datetime range for the selected date
    start_datetime = datetime.combine(selected_date, time.min)
    end_datetime = datetime.combine(selected_date, time.max)

    # Filter time entries that fall within the selected date
    qs = TimeEntry.objects.filter(
        time_in__gte=start_datetime,
        time_in__lte=end_datetime
    ).order_by("time_in")

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Entries"

    # Define headers for the Excel sheet
    headers = [
        "ID",
        "Employee ID",
        "First Name",
        "Surname",
        "Company",
        "Date",
        "Time In",
        "Time Out",
        "Hours Worked",
        "Is Late"
    ]
    ws.append(headers)

    # Populate rows with time entry data
    for entry in qs:
        row = [
            entry.id,
            entry.user.employee_id,
            entry.user.first_name,
            entry.user.surname,
            entry.user.company.name if entry.user.company else "",
            entry.time_in.strftime("%Y-%m-%d"),
            entry.time_in.strftime("%H:%M:%S"),
            entry.time_out.strftime("%H:%M:%S") if entry.time_out else "",
            entry.hours_worked,
            "Yes" if entry.is_late else "No"
        ]
        ws.append(row)

    # Save the workbook to an in-memory output buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create the response with a dynamic filename
    filename = f"time_entries_{date_str}.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"

    # Log the export action
    log_admin_action(request, "Excel Export", f"Exported {filename}")
    return response